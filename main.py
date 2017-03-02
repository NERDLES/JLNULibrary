#!/usr/bin/env python3 
# -*- coding: utf-8 -*- 

' 吉林师范大学图书馆学管会辅助排班程序 ' 

__author__ = '肖志强' 

import sys
import os
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

'''obj_'''
class Schedule(object):
	"""课表"""
	def __init__(self,table):
		super(Schedule, self).__init__()
		self.__table = []
		for i in range(1,8):
			for j in range(1,7):
				self.__table[i][j] = table[i][j]

"""学生信息"""
class Student(Schedule):
	free_course_list = set()
	def __init__(self, name,table):
		self.__dutytimes_res = 0
		self.__courses_count = 0
		self.__name = name
		self.__table=[]
		def calculate(self):
			for i in range(1,7):
				for j in range(1,8):
					if self.__table[i][j] != None:
						self.__courses_count += 1
					else:
						Student.free_course_list.add((i,j))
		for i in range(7):
			__app = []
			for j in range(8):
				__app.append(table[i][j])
			self.__table.append(__app)
		calculate(self)
	def get_name(self):
		return self.__name
	def get_dutytimes_res(self):
		return self.__dutytimes_res
	def set_dutytimes_res(self,dutytimes):
		self.__dutytimes_res = dutytimes
	def get_course_count(self):
		return self.__courses_count
	def get_table(self):
		return self.__table
	#self.__courses_count = courses_count
	#self.__dutytimes_res = dutytimes_res

class Course(object):
	"""统计每节课有哪些人没课"""
	def __init__(self, x, y):
		self.__member_list = []
		self.__member_count= 0
		self.__x = x
		self.__y = y
		self.__name = '空'
	def add_student(self,student):
		self.__member_list.append(student)
		self.__member_count = len(self.__member_list)
	def get_member_list(self):
		return self.__member_list
	def list_sort(self):
		pass
	def get_sta(self):
		return self.__x,self.__y
	def get_x(self):
		return self.__x
	def get_y(self):
		return self.__y
	def get_member_count(self):
		return self.__member_count
	def get_name(self):
		return self.__name
	def set_name(self,name):
		self.__name = name

class Mould_Schedule(Schedule):
	"""值班表模板"""
	__course_list = []
	duty_course_list = set()
	def __init__(self, table):
		__sta_x = 1
		__sta_y = 1
		for i in range(len(table)):
			for j in range(len(table[i])):
				if table[i][j].value == 'sta':
					__sta_x = i
					__sta_y = j
					break
		for i in range(__sta_x+1,len(table)):
			for j in range(__sta_y+1,len(table[i])):
				if table[i][j].value == 1:
					self.__course_list.append(Course(i-__sta_x,j-__sta_y))
					self.duty_course_list.add((i-__sta_x,j-__sta_y))
					#print 'duty ',(i-__sta_x,j-__sta_y)
	def get_course_list(self):
		return self.__course_list
	def set_course_list(self,stu):
		stu_table = stu.get_table()
		for i in range(len(stu_table)):
			for j in range(len(stu_table[i])):
				if stu_table[i][j] == None:
					for course in self.__course_list:
						if course.get_x() == i and course.get_y() == j:
							course.add_student(stu)
							break
	def sort_course_list(self):
		self.__course_list.sort(key = lambda Course : Course.get_member_count(),reverse = False)
	def write_name(self):
		__table = [
		['课程表' ,'星期一','星期二','星期三','星期四','星期五','星期六','星期日'],
		['1、2'	 ,''	 ,''	 ,''	  ,''	 ,''	  ,''	  ,''	  ],
		['3、4'	 ,''	 ,''	 ,''	  ,''	 ,''	  ,''	  ,''	  ],
		['5、6'	 ,''	 ,''	 ,''	  ,''	 ,''	  ,''	  ,''	  ],
		['7、8'	 ,''	 ,''	 ,''      ,''	 ,''	  ,''	  ,''	  ],
		['9、10'	 ,''	 ,''	 ,''	  ,''	 ,''	  ,''	  ,''     ],
		['11、12',''		 ,''	 ,''	  ,''	 ,''	  ,''	  ,''	  ],
		]
		for course in self.get_course_list():
			i = course.get_x()
			j = course.get_y()
			name = course.get_name()
			__table[i][j] = name
		return __table

	def print_table(self):
		for i in self.__table:
			print i
'''_obj'''

'''read&write_'''
#打开文件
def open_excel(file):
    try:
    	data = xlrd.open_workbook(file)
        return data
    except Exception,e:
        print str(e)

#获取Sheet1表格中的数据
def read_data(file):
	data  = load_workbook(filename = file)
	table = data['Sheet1']
	nrows = table.max_row #行数
	ncols = table.max_column #列数
	list  = []
	for i in table:
		app = []
		for j in i:
			app.append(j)

		list.append(app)
	return list
	
#写入文件并保存
def write_data(table):
	wb = Workbook()
	ws = wb.active
	ws.title = "Result"
	dest_filename = 'Rota.xlsx'
	for i in range(len(table)):
		for j in range(len(table[i])):
			ws.cell(row=i+1,column=j+1,value=table[i][j])
   	wb.save(filename = dest_filename)

def test():
	data = '/data.xlsx'
	path = os.getcwd()+data
	tables = read_data(path)
	write_data(tables)
'''_read&write'''

def main():
	data = '/data.xlsx'
	mould= '/mould.xlsx'
	data_path = os.getcwd()+data
	mould_path= os.getcwd()+mould
	data_tables = read_data(data_path)
	mould_table = read_data(mould_path)

	mould_schedule = Mould_Schedule(mould_table)
	stu_list = []

	'''创建学生表list'''
	for i in range(len(data_tables)):
		for j in range(len(data_tables[i])):
			if data_tables[i][j].value == 'Name':
				name = data_tables[i][j+1].value
				tmp_table = []
				for k in range(i+1,i+8):
					app = []
					for l in range(j,j+8):
						app.append(data_tables[k][l].value)
					tmp_table.append(app)
				stu_list.append(Student(name,tmp_table))

	#计算学生人数
	stunum_sum = len(stu_list)

	#计算需要值班的课程数目
	dutycourse_sum = len(Student.free_course_list & Mould_Schedule.duty_course_list)


	#计算人均值班次数
	dutytimes_avg = dutycourse_sum/stunum_sum
	dutytimes_res = dutycourse_sum%stunum_sum

	#按课程数分配值班数
	stu_list.sort(key = lambda Student : Student.get_course_count(),reverse = False)
	for i in range(stunum_sum):
		if i < dutytimes_res:
			stu_list[i].set_dutytimes_res(dutytimes_avg+1)
		else:
			stu_list[i].set_dutytimes_res(dutytimes_avg)

	'''
	print 'stunum =',stunum_sum,'coursenum =',dutycourse_sum
	print 'avg =',dutytimes_avg,'res =',dutytimes_res
	for i in stu_list:
		print 'name =',i.get_name(),'course_num =',i.get_course_count(),'dutytimes =',i.get_dutytimes_res()
	'''

	'''
	学生信息设置完毕
	'''

	#按学生课程数目降序排序
	stu_list.sort(key = lambda Student : Student.get_course_count(),reverse = True)

	#以学生课程数目为优先级顺序加入到MouldCourse中
	for stu in stu_list:
		mould_schedule.set_course_list(stu)

	#按照每节课所拥有空闲的学生数排序
	mould_schedule.sort_course_list()

	#按顺序填表
	for course in mould_schedule.get_course_list():
		for stu in course.get_member_list():
			if stu.get_dutytimes_res() > 0:
				course.set_name(stu.get_name())
				stu.set_dutytimes_res(stu.get_dutytimes_res()-1)
				break

	#将名字填入表中并保存
	write_data(mould_schedule.write_name())

if __name__=='__main__':
    main()

